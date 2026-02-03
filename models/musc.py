"""
MuSc: Zero-Shot Industrial Anomaly Classification and Segmentation.

This module implements the MuSc (Mutual Scoring) algorithm for zero-shot
anomaly detection, as described in the ICLR 2024 paper:

    "MuSc: Zero-Shot Industrial Anomaly Classification and Segmentation
     with Mutual Scoring of the Unlabeled Images"

The algorithm combines:
- LNAMD: Local Neighborhood Aggregation with Multi-scale Distance
- MSM: Mutual Scoring Module for comparing images without labels
- RsCIN: Reference-based Score Calibration for Image-level Normalization
"""

from __future__ import annotations

import os
import sys
import time
import warnings
from typing import Any, Dict, List, Optional, Tuple, Union

import cv2
import numpy as np
import torch
import torch.nn.functional as F
from numpy.typing import NDArray
from openpyxl import Workbook
from tqdm import tqdm

sys.path.append('./models/backbone')

import datasets.btad as btad
import datasets.mvtec as mvtec
import datasets.visa as visa
import models.backbone._backbones as _backbones
import models.backbone.open_clip as open_clip
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
from models.modules._LNAMD import LNAMD
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from utils.metrics import compute_metrics

warnings.filterwarnings("ignore")

# Type aliases
ConfigDict = Dict[str, Any]
ImageMetrics = Tuple[float, float, float]  # (AUROC, F1, AP)
PixelMetrics = Tuple[float, float, float, float]  # (AUROC, F1, AP, AuPRO)


class MuSc:
    """
    Zero-shot anomaly detection using mutual scoring of unlabeled images.

    This class provides the main interface for running MuSc anomaly detection.
    It supports multiple vision transformer backbones (DINO, DINOv2, CLIP, TIMM)
    and can process images in real-time or batch mode.

    Attributes:
        cfg: Configuration dictionary containing model and dataset settings.
        device: PyTorch device (cuda or cpu) for computation.
        model_name: Name of the backbone vision model.
        image_size: Input image resolution (e.g., 224, 384, 512).

    Example:
        >>> cfg = load_config("config.yaml")
        >>> model = MuSc(cfg)
        >>> # Real-time inference
        >>> anomaly_maps = model.infer_on_images([image_tensor])
        >>> # Batch evaluation on dataset
        >>> model.main()
    """

    def __init__(self, cfg: ConfigDict, seed: int = 0) -> None:
        """
        Initialize the MuSc model.

        Args:
            cfg: Configuration dictionary with the following structure:
                - datasets: img_resize, dataset_name, class_name, data_path, divide_num
                - models: backbone_name, batch_size, feature_layers, pretrained, r_list
                - testing: output_dir, vis, vis_type, save_excel
                - device: GPU index (int) or 'cpu'
            seed: Random seed for reproducibility.
        """
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device(
            "cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu"
        )

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()

    def load_backbone(self) -> None:
        """
        Load the vision transformer backbone model.

        Supports three types of backbones:
        - DINO/DINOv2: Self-supervised vision transformers from Meta
        - TIMM: Models from the timm library (e.g., vit_small_patch16_224)
        - CLIP/OpenCLIP: Contrastive language-image models

        The backbone is automatically selected based on the model_name prefix:
        - "dino_" or "dinov2_" → DINO/DINOv2
        - "vit_" → TIMM
        - Other → OpenCLIP

        Raises:
            RuntimeError: If the model cannot be loaded.
        """
        # 1. Handle DINO and DINOv2 models
        if self.model_name.startswith("dino_") or self.model_name.startswith("dinov2_"):
            self.dino_model = _backbones.load(self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None

        # 2. Handle TIMM models (new branch for models starting with "vit_")
        elif self.model_name.startswith("vit_"):
            import timm
            self.vision_model = timm.create_model(self.model_name, pretrained=True)
            self.vision_model.to(self.device)
            self.preprocess = None

        # 3. The remainder of your branches (CLIP, etc.)...
        else:
            # Your default / fallback code.
            import models.backbone.open_clip as open_clip
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(
                self.model_name, self.image_size, pretrained=self.pretrained
            )
            self.clip_model.to(self.device)
            self.vision_model = self.clip_model



    def load_datasets(
        self,
        category: str,
        divide_num: int = 1,
        divide_iter: int = 0,
    ) -> Any:
        """
        Load a test dataset for the specified category.

        Args:
            category: Product category name (e.g., 'bottle', 'screw').
            divide_num: Number of subdivisions for large datasets.
            divide_iter: Current subdivision index.

        Returns:
            Dataset object (MVTecDataset, VisaDataset, or BTADDataset).

        Raises:
            ValueError: If dataset_name is not recognized.
        """
        # dataloader
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=visa.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=btad.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset

    def visualization(
        self,
        image_path_list: List[str],
        gt_list: List[int],
        pr_px: NDArray[np.floating],
        category: str,
    ) -> None:
        """
        Save anomaly heatmap visualizations to disk.

        Args:
            image_path_list: List of source image file paths.
            gt_list: Ground truth labels (0=normal, 1=anomaly).
            pr_px: Predicted pixel-level anomaly maps [N, H, W].
            category: Category name for output subdirectory.
        """

        def normalization01(img: NDArray) -> NDArray:
            return (img - img.min()) / (img.max() - img.min())
        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map)*255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                save_path = os.path.join(self.output_dir, category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                save_path = os.path.join(save_path, img_name)
                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)

    def infer_on_images(
        self,
        images: List[torch.Tensor],
    ) -> NDArray[np.floating]:
        """
        Run anomaly detection inference on a batch of images.

        This is the main inference method for real-time and CLI usage.
        It processes images through the backbone, LNAMD, and MSM pipeline
        to produce pixel-level anomaly maps.

        Args:
            images: List of image tensors, each with shape [B, 3, H, W].
                   Typically a single tensor with batch dimension.

        Returns:
            Anomaly maps as numpy array with shape [N, H, W], where:
            - N is the number of images
            - H, W are the spatial dimensions (patch grid size)
            - Values are normalized to [0, 1] range

        Example:
            >>> tensor = torch.rand(1, 3, 224, 224).to(model.device)
            >>> anomaly_maps = model.infer_on_images([tensor])
            >>> max_score = anomaly_maps.max()
            >>> is_anomaly = max_score > 0.9
        """
        print("Running real-time inference on webcam frames…")

        # images comes in as a list; we want a single batch tensor [B,3,H,W]
        batch = images[0] if isinstance(images[0], torch.Tensor) else torch.cat(images)
        batch = batch.to(self.device)

        patch_tokens_list = []
        class_tokens = []

        with torch.no_grad(), torch.cuda.amp.autocast():
            # 1) DINO / DINOv2 branch
            if hasattr(self, 'dino_model'):
                patch_tokens_all = self.dino_model.get_intermediate_layers(batch, n=max(self.features_list))
                image_features   = self.dino_model(batch)
                patch_tokens     = [patch_tokens_all[l-1].cpu() for l in self.features_list]

            # 2) HuggingFace CLIP branch (using processor)
            elif hasattr(self, 'processor'):
                # prepare BGR→RGB numpy frames
                imgs = []
                for t in batch:
                    arr = (t.cpu().permute(1, 2, 0).numpy() * 255).astype('uint8')
                    imgs.append(cv2.cvtColor(arr, cv2.COLOR_RGB2BGR))
                inputs = self.processor(images=imgs, return_tensors="pt").to(self.device)
                # get patch embeddings from hidden states
                vision_out    = self.vision_model(inputs['pixel_values'], output_hidden_states=True)
                hidden_states = vision_out.hidden_states  # tuple of [B, seq_len, C]
                patch_tokens  = [hidden_states[l].cpu() for l in self.features_list]
                # get global features using CLIP's get_image_features method
                clip_out      = self.clip_model.get_image_features(**inputs)
                image_features = clip_out / clip_out.norm(dim=-1, keepdim=True)

            # 3) TIMM-based ViT branch (requires forward_features)
            elif hasattr(self, 'vision_model') and hasattr(self.vision_model, 'forward_features'):
                feats = self.vision_model.forward_features(batch)
                # If the features are spatial (4D tensor), flatten to [B, tokens, C]
                if feats.ndim == 4:
                    B, C, H, W = feats.shape
                    feats = feats.reshape(B, C, H * W).permute(0, 2, 1)
                global_feats   = feats.mean(dim=1)
                image_features = global_feats / global_feats.norm(dim=-1, keepdim=True)
                patch_tokens   = [feats.cpu()]

            # 4) (Optional) Additional branch if you use a separate vit_model
            elif hasattr(self, 'vit_model'):
                feats = self.vit_model.forward_features(batch)
                if feats.ndim == 4:
                    B, C, H, W = feats.shape
                    feats = feats.reshape(B, C, H * W).permute(0, 2, 1)
                global_feats   = feats.mean(dim=1)
                image_features = global_feats / global_feats.norm(dim=-1, keepdim=True)
                patch_tokens   = [feats.cpu()]

            # 5) open_clip fallback branch
            else:
                image_features, patch_tokens_all = self.clip_model.encode_image(batch, self.features_list)
                image_features /= image_features.norm(dim=-1, keepdim=True)
                if not patch_tokens_all:
                    raise ValueError("ERROR: patch_tokens is empty! The model did not extract any features.")
                # Adjust if fewer layers than expected:
                num_avail = len(patch_tokens_all)
                if num_avail < len(self.features_list):
                    self.features_list = self.features_list[:num_avail]
                patch_tokens = [patch_tokens_all[l].cpu() for l in range(len(self.features_list))]

        # Stash the global features for any classification or scoring
        class_tokens.extend(
            [image_features[i].squeeze().cpu().numpy() for i in range(image_features.shape[0])]
        )
        patch_tokens_list.append(patch_tokens)

        # … then continue with your LNAMD + MSM pipeline as before …



        # Run LNAMD and MSM operations
        print("Processing features with LNAMD and MSM...")
        feature_dim = patch_tokens_list[0][0].shape[-1]
        anomaly_maps_r = torch.tensor([]).double()

        for r in self.r_list:
            LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
            # Embed and normalize
            features = LNAMD_r._embed([p.to(self.device) for p in patch_tokens_list[0]])
            features /= features.norm(dim=-1, keepdim=True)

            print(f"Shape of Z before MSM: {features.shape}")  # Debugging

            # Ensure Z has exactly 3 dimensions: [batch_size, num_patches, feature_dim]
            features = features.squeeze(2)  # remove extra dim if present
            if features.ndim != 3:
                raise ValueError(f"Invalid shape for Z: {features.shape}. Expected 3D tensor.")

            anomaly_maps_msm = MSM(Z=features, device=self.device, topmin_min=0, topmin_max=0.3)
            # (Repeat MSM if needed — your code calls MSM twice but that might be a duplicate)
            anomaly_maps_msm = MSM(Z=features, device=self.device, topmin_min=0, topmin_max=0.3)
            
            anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)

        # Average across the r_list dimension
        anomaly_maps = torch.mean(anomaly_maps_r, 0).cpu().numpy()
        print("Anomaly detection complete. Visualizing results...")

        # NOTE: At this point, anomaly_maps is typically [N, patch_count].
        # We need to reshape it to [N, h, w] if patch_count = h*w.

        pr_px = anomaly_maps
        # Attempt to reshape if needed:
        #   - If [N, patch_count], reshape to [N, side, side].
        #   - If [N, 1, H, W], drop the channel dim.
        #   - If it's already [N, H, W], do nothing.
        if pr_px.ndim == 2:
            # shape = [N, patch_count]
            side_length = int(np.sqrt(pr_px.shape[1]))
            pr_px = pr_px.reshape(pr_px.shape[0], side_length, side_length)
        elif pr_px.ndim == 4 and pr_px.shape[1] == 1:
            # shape = [N, 1, H, W]
            pr_px = pr_px[:, 0, :, :]
        elif pr_px.ndim not in [3]:
            raise ValueError(f"Unexpected anomaly_maps shape: {pr_px.shape}")

        # Now pr_px is [N, H, W]. Then we find the maximum score per image:
        max_scores = pr_px.reshape(pr_px.shape[0], -1).max(axis=1)
        anomaly_threshold = 0.99
        high_anomaly_indices = np.where(max_scores > anomaly_threshold)[0]

        if len(high_anomaly_indices) > 0:
            for i in high_anomaly_indices:
                # Extract single anomaly map: shape [H, W]
                anomaly_map = pr_px[i]

                # Normalize to [0, 255]
                anomaly_map = (anomaly_map - anomaly_map.min()) / (anomaly_map.max() - anomaly_map.min() + 1e-8)
                anomaly_map = (anomaly_map * 255).astype(np.uint8)

                # Convert single-channel to color
                anomaly_map = cv2.applyColorMap(anomaly_map, cv2.COLORMAP_JET)

                # Reconstruct the original frame
                # (images[i] is in [C,H,W], scale it back to 0..255 in BGR for cv2)
                frame = batch[i].cpu().numpy().transpose(1, 2, 0) * 255
                frame = frame.astype(np.uint8)
                frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)

                # Resize anomaly_map if it doesn't match frame size
                if anomaly_map.shape[:2] != frame.shape[:2]:
                    anomaly_map = cv2.resize(anomaly_map, (frame.shape[1], frame.shape[0]))

                # (Optional) Overlay the anomaly_map on top of the frame, etc.
                # e.g. alpha-blend or just display it

                cv2.waitKey(1)

        # Optionally, simply clip the output without dividing by the batch maximum
        pr_px = pr_px.astype(np.float32)
        pr_px = np.clip(pr_px, 0.0, 1.0)



        return pr_px

    def make_category_data(self, category: str) -> Tuple[ImageMetrics, PixelMetrics]:
        """
        Process all test images for a single category and compute metrics.

        This method runs the full MuSc pipeline on a dataset category:
        1. Load test dataset
        2. Extract features using the backbone
        3. Apply LNAMD for feature aggregation
        4. Apply MSM for mutual scoring
        5. Apply RsCIN for image-level score calibration
        6. Compute evaluation metrics

        Args:
            category: Product category name (e.g., 'bottle', 'screw').

        Returns:
            Tuple of (image_metrics, pixel_metrics) where:
            - image_metrics: (AUROC, F1, AP) for image-level classification
            - pixel_metrics: (AUROC, F1, AP, AuPRO) for pixel-level segmentation
        """
        print(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)
            test_dataloader = torch.utils.data.DataLoader(
                test_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )
            
            # extract features
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num
            start_time = time.time()
            for image_info in tqdm(test_dataloader):
            # for image_info in test_dataloader:
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if 'dinov2' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                    else: # clip
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                        image_features /= image_features.norm(dim=-1, keepdim=True)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]
                class_tokens.extend(image_features)
                patch_tokens_list.append(patch_tokens)  # (B, L+1, C)
            end_time = time.time()
            print('extract time: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            
            # LNAMD
            feature_dim = patch_tokens_list[0][0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double()
            for r in self.r_list:
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)
                        features /= features.norm(dim=-1, keepdim=True)
                        for l in range(len(self.features_list)):
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time-start_time)*1000/subset_num))

                # MSM
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device) # (N, L, C)
                    print('layer-{} mutual scoring...'.format(l))
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3)
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)
                    torch.cuda.empty_cache()
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)
            del anomaly_maps_r
            torch.cuda.empty_cache()

            # interpolate
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.cpu()), dim=0)
            anomaly_maps = anomaly_maps.cpu().numpy()

            # normalize so 1.0 is the top score
            anomaly_maps = anomaly_maps.astype(np.float32)
            global_max = anomaly_maps.max()
            if global_max > 0:
                anomaly_maps = anomaly_maps / global_max
            anomaly_maps = np.clip(anomaly_maps, 0.0, 1.0)

            B = anomaly_maps.shape[0]
            ac_score = anomaly_maps.reshape(B, -1).max(-1)

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        print('MuSc: {}ms per image'.format((end_time_all-start_time_all)*1000/dataset_num))

        anomaly_maps = anomaly_maps.cpu().numpy()
        torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        # RsCIN
        if self.dataset == 'visa':
            k_score = [1, 8, 9]
        elif self.dataset == 'mvtec_ad':
            k_score = [1, 2, 3]
        else:
            k_score = [1, 2, 3]
        scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)

        print('computing metrics...')
        pr_sp = np.array(scores_cls)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)
    
        return image_metric, pixel_metric

    def main(self) -> None:
        """
        Run full benchmark evaluation on all configured categories.

        Processes each category in self.categories, computes metrics,
        prints results, and optionally saves to Excel file.

        Results include:
        - Per-category image-level metrics (AUROC, F1, AP)
        - Per-category pixel-level metrics (AUROC, F1, AP, AuPRO)
        - Mean metrics across all categories
        """
        auroc_sp_ls: List[float] = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))


